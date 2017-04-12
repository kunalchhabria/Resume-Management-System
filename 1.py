import openpyxl,smtplib,sys,imaplib,pyzmail,email,pprint,os,re,PyPDF2,docx
pat = re.compile(r'\s+')
pat1=re.compile(r'\n+')
received_from_data={}
#received_from_data[id]=[id,name,date,filepath,number,decision]
print(' checking emails  with subject \'RESUME \' and are unread.. \n enter email id where the applicants have sent resumes')
receiveId=input()
print('enter password')
receivePass=input()

print('Specify the skills you are looking for in resumes. Enter all the skill seperated by spaces. \nexample: python django java')
skillset=input().split(' ')
print('\n'*1000)
def inboxSearch():
    print('Searching for the resumes...\n\n')
    detach_dir = 'c:/python36/pyt/mails'
    m = imaplib.IMAP4_SSL("imap.gmail.com")
    
    
    m.login(receiveId,receivePass)
    m.select("inbox")

    resp, items = m.search(None,'(SUBJECT "resume")',)
    items = items[0].split()
    #print(resp,items)
    print(len(items))
    for emailid in items:
        resp, data = m.fetch(emailid, "(RFC822)") 
        email_body = data[0][1] 
        mail = email.message_from_bytes(email_body) 
        temp = m.store(emailid,'+FLAGS', '\\Seen')
        m.expunge()

        if mail.get_content_maintype() != 'multipart':
            continue

        received_from=mail["From"]
        email_start_index=received_from.index('<')+1
        email_end_index=received_from.index('>')
        received_from_emailid=received_from[email_start_index:email_end_index]
        received_from_name=received_from[:email_start_index-1]
        received_from_date=mail["Date"]
        
        print ("["+mail["From"]+"] :" + mail["Subject"])

        for part in mail.walk():
            if part.get_content_maintype() == 'multipart':
                continue
            if part.get('Content-Disposition') is None:
                continue
            if part.get_filename().endswith('.pdf'):
                file_type='.pdf'
            if part.get_filename().endswith('.docx'):
                file_type='.docx'
            
            filename = received_from_emailid+file_type
            att_path = os.path.join(detach_dir, filename)

            if not os.path.isfile(att_path) :
                fp = open(att_path, 'wb')
                fp.write(part.get_payload(decode=True))
                fp.close()
                received_from_data[received_from_emailid]=[received_from_emailid,received_from_name,received_from_date,att_path]
    print(received_from_data)
    print ('Finished downloading resumes.\n\n')

def extractText():
	print('Scanning all the resumes...\n\n')
	pat = re.compile(r'\s+')
	pat1=re.compile(r'\n+')
	pat2=re.compile(r'\d{10,12}')
	pat3=re.compile(r'\d{3}[-,\s]\d{3}[-,\s]\d{4}')
	for downloaded_resume in received_from_data:
		content=''
		if received_from_data[downloaded_resume][3].endswith('.pdf'):
			path=received_from_data[downloaded_resume][3]
			pdfFileObj = open(path, 'rb')
			pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
			content=''
			pages=pdfReader.numPages
			
			for i in range(pages):
				pageObj = pdfReader.getPage(i)
				content+=pageObj.extractText()
			content=pat.sub('',content)
			content=pat1.sub('',content)


		if received_from_data[downloaded_resume][3].endswith('.docx'):
			path=received_from_data[downloaded_resume][3]
			doc = docx.Document(path)
			content=''
			for para in doc.paragraphs:
				content.append(para.text)
			content==''.join(para.text)
			content=pat.sub('',content)
			content=pat1.sub('',content)

		phones_numbers1=pat2.findall(content)
		phones_numbers2=pat3.findall(content)
		phones_numbers_all=','.join(phones_numbers1) + ','.join(phones_numbers2)			
		if len(phones_numbers_all)<2:
			received_from_data[downloaded_resume].append('-')
		else:
			received_from_data[downloaded_resume].append(phones_numbers_all)

		flag=True
		for required_skill in skillset:
			if not required_skill in content:
				flag=False
		if flag is True:
			received_from_data[downloaded_resume].append('Yes')
		else:
			received_from_data[downloaded_resume].append('No')
	print(received_from_data)
	print('Finished scanning all the resumes.\n\n')




def saveInXl():
	#received_from_data[id]=[id,name,date,filepath,number,decision]
	print('Saving data in excel sheet...\n')
	wb=openpyxl.Workbook()
	sheet=wb.active
	sheet.title='resumes'
	sheet.cell(row=1,column=1).value='NAME'
	sheet.cell(row=1,column=2).value='PHONE NUMBER'
	sheet.cell(row=1,column=3).value='EMAIL ID'
	sheet.cell(row=1,column=4).value='DATE-TIME'
	sheet.cell(row=1,column=5).value='DECISION'

	sheet_row=2
	for downloaded_resume in received_from_data:
		sheet.cell(row=sheet_row,column=1).value=received_from_data[downloaded_resume][1]
		sheet.cell(row=sheet_row,column=2).value=received_from_data[downloaded_resume][4]
		sheet.cell(row=sheet_row,column=3).value=received_from_data[downloaded_resume][0]
		sheet.cell(row=sheet_row,column=4).value=received_from_data[downloaded_resume][2]
		sheet.cell(row=sheet_row,column=5).value=received_from_data[downloaded_resume][5]
		sheet_row+=1



	'''
	for i in range(2,10):
		for j in range(1,6):
			if j==4:
				temp=' '
				for k in results[3]:
					temp+= k+','
				temp=temp[:(len(temp)-1)]
				sheet.cell(row=i,column=j).value=temp
			else:
				sheet.cell(row=i,column=j).value=results[j-1]
		if i%2==0:
			sheet.cell(row=i,column=6).value='yes'
		else:
			sheet.cell(row=i,column=6).value='no'
	'''
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 40
	sheet.column_dimensions['C'].width = 40
	sheet.column_dimensions['D'].width = 40
	sheet.column_dimensions['E'].width = 20
	wb.save('resume.xlsx')
	print("Finished saving data in excel sheet.\n\n")


def sendmail():
	print("Sending replies to candidates...\n ")
	smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
	smtpObj.ehlo()
	smtpObj.starttls()
	#print('enter email id')
	useremail=receiveId
	#print('enter password')
	password=receivePass
	smtpObj.login(useremail,password)
	wb=openpyxl.load_workbook('resume.xlsx')
	sheet=wb.get_sheet_by_name('resumes')
	lastCol=5

	for r in range(2, 2+len(received_from_data)):
		decision=sheet.cell(row=r, column=lastCol).value
		name=sheet.cell(row=r, column=1).value
		senderemail=sheet.cell(row=r, column=3).value
		if decision=='Yes':
			body = "Subject: SELECTED.\nDear %s,\n We are glad to inform you that you are selected for techincal interview." %(name)

		else:
			body = "Subject: rejected.\nDear %s,\n We are sorry to inform you that you are not selected for techincal interview." %(name)
		print('Sending email to %s...' % senderemail)
		sendmailStatus = smtpObj.sendmail(useremail, senderemail, body)

		if sendmailStatus != {}:
			print('There was a problem sending email to %s: %s' % (sendereemail,sendmailStatus))


	smtpObj.quit()
	print("Finished sending replies to candidates.\n\n")


inboxSearch()
extractText()
saveInXl()
sendmail()

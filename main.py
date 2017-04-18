import openpyxl,smtplib,sys,imaplib,pyzmail,email,pprint,os,re,PyPDF2,docx,datetime
pat = re.compile(r'\s+')
pat1=re.compile(r'\n+')
received_from_data={}

detach_dir = 'resumes-and-candidate-data\\'
now_time=str(datetime.datetime.now())
micro_second_index=now_time.index('.')
now_time=now_time[:micro_second_index]
detach_dir=detach_dir+now_time
detach_dir=detach_dir.replace(' ',',')
detach_dir=detach_dir.replace(':','-')
if not os.path.exists(detach_dir):
	os.makedirs(detach_dir)

#received_from_data[id]=[id,name,date,filepath,number,decision]
print(' checking emails  with subject \'RESUME \' and are unread.. \n enter email id where the applicants have sent resumes')
receiveId=input()
print('enter password')
receivePass=input()

print('Specify the skills you are looking for in resumes. Enter all the skill seperated by spaces. \nexample: python django java')
skillset=input().split(' ')
print('\n'*1000)
print(detach_dir)
def inboxSearch():
    print('Searching for the resumes...\n\n')
    
    m = imaplib.IMAP4_SSL("imap.gmail.com")
    
    '''if not os.path.exists(directory):
    os.makedirs(directory)
    '''
    m.login(receiveId,receivePass)
    m.select("inbox")

    
    #print(resp,items)

    resp, items = m.search(None,'(UNSEEN SUBJECT "resume")',)
    items = items[0].split()
    #print(items,len(items))

    while(len(items)>0):
	    try:
	        emailid=items[len(items)-1]
	        resp, data = m.fetch(emailid, "(RFC822)") 
	        
	        email_body = data[0][1]
	        email_body=email_body.decode('utf-8')
	        mail = email.message_from_string(email_body) 
	        temp = m.store(emailid,'+FLAGS', '\\Seen')
	        m.expunge()
	        removed=items.pop()

	        if mail.get_content_maintype() != 'multipart':
	            continue

	        received_from=mail["From"]
	        email_start_index=received_from.index('<')+1
	        email_end_index=received_from.index('>')
	        received_from_emailid=received_from[email_start_index:email_end_index]
	        received_from_name=received_from[:email_start_index-1]
	        received_from_date=mail["Date"]
	        
	        print ("["+mail["From"]+"] :" + mail["Subject"])

	        for part in mail.walk():
	            if part.get_content_maintype() == 'multipart':
	                continue
	            if part.get('Content-Disposition') is None:
	                continue
	            if part.get_filename().endswith('.pdf'):
	                file_type='.pdf'
	            if part.get_filename().endswith('.docx'):
	                file_type='.docx'
	            
	            filename = received_from_emailid+file_type
	            att_path = os.path.join(detach_dir, filename)

	            if not os.path.isfile(att_path) :
	                fp = open(att_path, 'wb')
	                fp.write(part.get_payload(decode=True))
	                fp.close()
	                received_from_data[received_from_emailid]=[received_from_emailid,received_from_name,received_from_date,att_path]
	    except:
	    	asdf=1 #do nothing
    #print(received_from_data)
    print ('Finished downloading resumes.\n\n')

def extractText():
	print('Scanning all the resumes...\n\n')
	pat = re.compile(r'\s+')
	pat1=re.compile(r'\n+')
	pat2=re.compile(r'\d{10,12}')
	pat3=re.compile(r'\d{3}[-,\s]\d{3}[-,\s]\d{4}')
	for downloaded_resume in received_from_data:
		content=''
		if received_from_data[downloaded_resume][3].endswith('.pdf'):
			path=received_from_data[downloaded_resume][3]
			pdfFileObj = open(path, 'rb')
			pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
			content=''
			pages=pdfReader.numPages
			
			for i in range(pages):
				pageObj = pdfReader.getPage(i)
				content+=pageObj.extractText()
			content=pat.sub('',content)
			content=pat1.sub('',content)


		if received_from_data[downloaded_resume][3].endswith('.docx'):
			path=received_from_data[downloaded_resume][3]
			doc = docx.Document(path)
			content=[]
			for para in doc.paragraphs:
				content.append(para.text)
			content=''.join(para.text)
			content=pat.sub('',content)
			content=pat1.sub('',content)

		content=content.lower()
		#print('length ',len(content))
		phones_numbers1=pat2.findall(content)
		phones_numbers2=pat3.findall(content)
		phones_numbers_all=','.join(phones_numbers1) + ','.join(phones_numbers2)			
		if len(phones_numbers_all)<2:
			received_from_data[downloaded_resume].append('-')
		else:
			received_from_data[downloaded_resume].append(phones_numbers_all)

		flag=True
		for required_skill in skillset:
			if not required_skill in content:
				flag=False
		if flag is True:
			received_from_data[downloaded_resume].append('Yes')
		else:
			received_from_data[downloaded_resume].append('No')
	#print(received_from_data)
	print('Finished scanning all the resumes.\n\n')




def saveInXl():
	#received_from_data[id]=[id,name,date,filepath,number,decision]
	print('Saving data in excel sheet...\n')
	wb=openpyxl.Workbook()
	sheet=wb.active
	sheet.title='resumes'
	sheet.cell(row=1,column=1).value='NAME'
	sheet.cell(row=1,column=2).value='PHONE NUMBER'
	sheet.cell(row=1,column=3).value='EMAIL ID'
	sheet.cell(row=1,column=4).value='DATE-TIME'
	sheet.cell(row=1,column=5).value='DECISION'

	sheet_row=2
	for downloaded_resume in received_from_data:
		sheet.cell(row=sheet_row,column=1).value=received_from_data[downloaded_resume][1]
		sheet.cell(row=sheet_row,column=2).value=received_from_data[downloaded_resume][4]
		sheet.cell(row=sheet_row,column=3).value=received_from_data[downloaded_resume][0]
		sheet.cell(row=sheet_row,column=4).value=received_from_data[downloaded_resume][2]
		sheet.cell(row=sheet_row,column=5).value=received_from_data[downloaded_resume][5]
		sheet_row+=1


	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 40
	sheet.column_dimensions['C'].width = 40
	sheet.column_dimensions['D'].width = 40
	sheet.column_dimensions['E'].width = 20
	wb.save(detach_dir+ '\\candidate_data.xlsx')
	print("Finished saving data in excel sheet.\n\n")


def sendmail():
	print("Sending replies to candidates...\n ")
	smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
	smtpObj.ehlo()
	smtpObj.starttls()
	#print('enter email id')
	useremail=receiveId
	#print('enter password')
	password=receivePass
	smtpObj.login(useremail,password)
	wb=openpyxl.load_workbook(detach_dir+ '\\candidate_data.xlsx')
	sheet=wb.get_sheet_by_name('resumes')
	lastCol=5

	for r in range(2, 2+len(received_from_data)):
		decision=sheet.cell(row=r, column=lastCol).value
		name=sheet.cell(row=r, column=1).value
		senderemail=sheet.cell(row=r, column=3).value
		if decision=='Yes':
			body = "Subject: SELECTED.\nDear %s,\n We are glad to inform you that you are selected for techincal interview." %(name)

		else:
			body = "Subject: rejected.\nDear %s,\n We are sorry to inform you that you are not selected for techincal interview." %(name)
		print('Sending email to %s...' % senderemail)
		sendmailStatus = smtpObj.sendmail(useremail, senderemail, body)

		if sendmailStatus != {}:
			print('There was a problem sending email to %s: %s' % (sendereemail,sendmailStatus))


	smtpObj.quit()
	print("Finished sending replies to candidates.\n\n")


inboxSearch()
extractText()
saveInXl()
sendmail()
input()
